import copy
import json
from pathlib import Path

from openpyxl import load_workbook

sample = {
    'weapon': {'5': [], '4': [], '3': []},
    'artifact': [],  # list[list[str,Optional[str]]]  四件套 / 2+2
    'remark': [],
}

char_json = {}

path = Path(__file__).parents[1] / 'genshinuid_adv'
data_path = Path(__file__).parent / 'help_data'

wb = load_workbook(str(data_path / 'Genshin All Char.xlsx'))
ws = wb.active
for char_i in range(2, 1700, 5):  # 角色行
    char = ws.cell(char_i, 1).value
    if char is None:
        for _i in range(5):
            val = ws.cell(char_i + _i, 1).value
            if val is not None:
                char = val
    if not isinstance(char, str):
        continue
    char_name = char.replace('\n', '')
    char_sample = copy.deepcopy(sample)
    for i in range(5):
        row = i + char_i

        if star_5 := ws.cell(row, 2).value:
            char_sample['weapon']['5'].append(star_5)
        if star_4 := ws.cell(row, 3).value:
            char_sample['weapon']['4'].append(star_4)
        if star_3 := ws.cell(row, 4).value:
            char_sample['weapon']['3'].append(star_3)

        artifact = []
        if arti_1 := ws.cell(row, 5).value:
            artifact.append(arti_1)
        if arti_2 := ws.cell(row, 6).value:
            artifact.append(arti_2)
        if artifact:
            char_sample['artifact'].append(artifact)

        if remark := ws.cell(row, 7).value:
            if row > 7:
                char_sample['remark'].append(remark)

    char_json[char_name] = char_sample

with open(str(path / 'char_adv_list.json'), 'w', encoding='utf-8') as f:
    json.dump(char_json, f, indent=2, ensure_ascii=False)
