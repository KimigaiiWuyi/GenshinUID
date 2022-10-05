import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

R_PATH = Path(__file__).parent
DATA_PATH = R_PATH / 'blue_data'
CURVE_PATH = Path(__file__).parents[1] / 'genshinuid_enka' / 'curveCalc'

wb = load_workbook(str(DATA_PATH / '曲线素材.xlsx'), data_only=True)
ws: Worksheet = wb['曲线素材']  # type: ignore

result = {}
for col in range(1, 45):
    if not ws.cell(1, col).value:
        continue
    else:
        title = ws.cell(1, col).column_letter  # type: ignore

    temp = []
    for row in range(2, 302):
        val = ws.cell(row, col).value
        temp.append(val)

    result[title] = temp

with open(str(CURVE_PATH / 'curve.json'), 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
