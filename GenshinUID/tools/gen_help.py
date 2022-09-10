import copy
import json
from pathlib import Path

from openpyxl import load_workbook

sample = {
    'name': '',
    'desc': '',
    'eg': '',
    'need_ck': False,
    'need_sk': False,
    'need_admin': False,
}

result = {}

R_PATH = Path(__file__).parent
DATA_PATH = R_PATH / 'help_data'
HELP_PATH = Path(__file__).parents[1] / 'genshinuid_help'

# 表格地址: https://kdocs.cn/l/ccpc6z0bZx6u
wb = load_workbook(str(DATA_PATH / 'GenshinUID Help.xlsx'))
ws = wb.active

module_name_str = ''
for row in range(2, 999):
    # 跳过空白行
    if not ws.cell(row, 2).value:
        continue

    _sample = copy.deepcopy(sample)

    # 将第一列读取为模块名
    if ws.cell(row, 1):
        module_name_str = ws.cell(row, 1).value

    # 第二列为功能名
    _sample['name'] = ws.cell(row, 2).value
    # 第三列为详细信息
    _sample['desc'] = ws.cell(row, 3).value
    # 第四列为使用例
    _sample['eg'] = ws.cell(row, 4).value

    if ws.cell(row, 5).value == '是':
        _sample['need_ck'] = True

    if ws.cell(row, 6).value == '是':
        _sample['need_sk'] = True

    if ws.cell(row, 7).value == '是':
        _sample['need_admin'] = True

    if module_name_str is None:
        pass
    else:
        module_name = module_name_str.split(' | ')[0]  # type: ignore
        module_desc = module_name_str.split(' | ')[1]  # type: ignore
        if module_name not in result:
            result[module_name] = {'desc': module_desc, 'data': []}

    result[module_name]['data'].append(_sample)  # type: ignore

with open(str(HELP_PATH / 'help.json'), 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
