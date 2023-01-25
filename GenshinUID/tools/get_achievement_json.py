import json
import warnings
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import_path = Path(__file__).parent / 'achievement_data' / '成就汇总.xlsx'
export_path = Path(__file__).parents[1] / 'genshinuid_achievement'

wb: Workbook = load_workbook(str(import_path))
ws_daily: Worksheet = wb['成就相关每日委托']
ws_all: Worksheet = wb['正式服成就汇总']
ws_32: Worksheet = wb['3.2 新增成就']
ws_33: Worksheet = wb['3.3 新增成就']
ws_34: Worksheet = wb['3.4 新增成就']

result_achi = {}
result_all = {}

is_first = False
for row in range(3, 100):
    task = ws_daily.cell(row, 3).value
    achi = ws_daily.cell(row, 4).value
    desc = ws_daily.cell(row, 5).value
    guide = ws_daily.cell(row, 6).value
    hyper_link = ws_daily.cell(row, 6).hyperlink
    hyper_link = '' if hyper_link is None else hyper_link.target
    if not task:
        if is_first:
            break
        is_first = True
        continue
    else:
        is_first = False
    task_list = task.split('\n')
    for t in task_list:
        if t.startswith('('):
            continue
        result_achi[t] = {
            'achievement': achi,
            'desc': desc,
            'guide': guide,
            'link': hyper_link,
        }

for row in range(3, 1000):
    book = ws_all.cell(row, 5).value
    achi = ws_all.cell(row, 6).value
    desc = ws_all.cell(row, 7).value
    guide = ws_all.cell(row, 11).value
    hyper_link = ws_all.cell(row, 11).hyperlink
    hyper_link = '' if hyper_link is None else hyper_link.target
    if not book:
        break
    result_all[achi] = {
        'book': book,
        'desc': desc,
        'guide': guide,
        'link': hyper_link,
    }


def get_book(_book: Worksheet, loop: int, bn: int, an: int, dn: int, gn: int):
    for row in range(3, loop):
        book = _book.cell(row, bn).value
        achi = _book.cell(row, an).value
        desc = _book.cell(row, dn).value
        guide = _book.cell(row, gn).value
        hyper_link = _book.cell(row, gn).hyperlink
        hyper_link = '' if hyper_link is None else hyper_link.target
        if not book:
            break
        result_all[achi] = {
            'book': book,
            'desc': desc,
            'guide': guide,
            'link': hyper_link,
        }


get_book(ws_32, 50, 4, 5, 6, 9)
get_book(ws_33, 50, 5, 6, 7, 10)
get_book(ws_34, 50, 5, 6, 7, 10)


with open(str(export_path / 'daily_achi.json'), 'w', encoding='utf-8') as f:
    json.dump(result_achi, f, indent=2, ensure_ascii=False)

with open(str(export_path / 'all_achi.json'), 'w', encoding='utf-8') as f:
    json.dump(result_all, f, indent=2, ensure_ascii=False)
