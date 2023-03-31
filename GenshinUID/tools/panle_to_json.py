import json
import asyncio
from pathlib import Path

import httpx
import openpyxl

R_PATH = Path(__file__).parent
DATA_PATH = R_PATH / 'blue_data'
ETC_PATH = Path(__file__).parents[1] / 'genshinuid_enka' / 'effect'


SAMPLE = {
    '神里绫华': 'Q切割伤害',
    '琴': 'Q领域发动治疗量',
    '丽莎': ['E三层引雷长按伤害', 'Q放电伤害'],
    '芭芭拉': 'Q治疗量',
    '凯亚': 'Q伤害',
    '迪卢克': 'Q斩击伤害(蒸发)',
    '雷泽': 'A一段伤害',
    '安柏': ['Q箭雨总伤害', 'A满蓄力瞄准射击'],
    '温迪': 'A扩散伤害',
    '香菱': 'Q旋火轮伤害(蒸发)',
    '北斗': 'Q闪雷伤害',
    '行秋': 'Q剑雨伤害',
    '魈': 'A高空下落伤害',
    '凝光': 'Q宝石伤害',
    '可莉': 'A重击伤害(蒸发)',
    '钟离': ['E总护盾量', 'Q伤害'],
    '菲谢尔': 'E奥兹攻击伤害',
    '班尼特': ['Q伤害(蒸发)', 'Q持续治疗'],
    '达达利亚': 'Q伤害·近战(蒸发)',
    '诺艾尔': 'A一段伤害',
    '七七': 'Q治疗量',
    '重云': 'Q伤害',
    '甘雨': ['A霜华矢·霜华绽发伤害', 'A霜华矢·霜华绽发伤害(融化)'],
    '阿贝多': 'E刹那之花伤害',
    '迪奥娜': 'E护盾基础吸收量',
    '莫娜': ['Q泡影破裂伤害(蒸发)', 'Q泡影破裂伤害'],
    '刻晴': 'A重击伤害',
    '砂糖': 'A扩散伤害',
    '辛焱': ['Q伤害', 'E三级护盾吸收量'],
    '罗莎莉亚': 'Q冰枪持续伤害',
    '胡桃': 'A重击伤害(蒸发)',
    '枫原万叶': 'A扩散伤害',
    '烟绯': 'A重击伤害(蒸发)',
    '宵宫': 'A一段伤害',
    '托马': 'E护盾吸收量上限',
    '优菈': 'Q光降之剑基础伤害(13层)',
    '雷电将军': 'Q梦想一刀基础伤害(满愿力)',
    '早柚': 'Q不倒貉貉治疗量',
    '珊瑚宫心海': ['A一段伤害', 'E治疗量'],
    '五郎': 'def',
    '九条裟罗': 'Q天狗咒雷·金刚坏 伤害',
    '荒泷一斗': 'A荒泷逆袈裟连斩伤害',
    '八重神子': 'E杀生樱伤害·叁阶',
    '鹿野院平藏': 'E伤害',
    '夜兰': 'Q玄掷玲珑伤害',
    '埃洛伊': 'Q伤害',
    '申鹤': 'atk',
    '云堇': 'def',
    '久岐忍': ['E越祓草轮治疗量', 'A元素反应(超绽放)'],
    '神里绫人': 'E一段瞬水剑伤害',
    '柯莱': 'Q跃动伤害(蔓激化)',
    '多莉': 'Q持续治疗量',
    '提纳里': 'A藏蕴花矢伤害(蔓激化)',
    '妮露': 'A丰穰之核(绽放)',
    '赛诺': 'E渡荒之雷(超激化)',
    '坎蒂丝': 'Q水波冲击伤害',
    '纳西妲': 'E灭净三业伤害(蔓激化)',
    '莱依拉': 'E护盾基础吸收量',
    '旅行者(草)': 'Q草灯莲攻击伤害(蔓激化)',
    '旅行者(岩)': 'Q地震波单次伤害',
    '旅行者(风)': 'A扩散伤害',
    '流浪者': 'E空居·不生断伤害',
    '珐露珊': 'Q伤害',
    '艾尔海森': 'E1枚光幕攻击伤害(蔓激化)',
    '瑶瑶': 'E白玉萝卜治疗量',
    '米卡': 'Q施放治疗量',
}


async def get_misc_info(mode: str, name: str):
    '''
    :说明:
      一些杂项信息。
    :参数:
      * name (str): 'enemies', 'foods', 'artifacts'。
      * name (str): 参数。
    :返回:
      * data (str): 获取数据信息。
    '''
    url = 'https://info.minigg.cn/{}'.format(mode)
    req = httpx.get(url=url, params={'query': name})
    return req.json()


async def getEquipName(name: str) -> str:
    if name == '花神':
        name = '乐园遗落之花'
    elif name == '饰金':
        name = '饰金之梦'
    elif name == '追忆':
        name = '追忆之注连'
    print(name)
    r = await get_misc_info('artifacts', name)
    re = r['name']
    print(re)
    return re


async def panle2Json() -> None:
    '''
    :说明:
      访问DATA_PATH并转换数据为dmgMap.json。
    '''
    wb = openpyxl.load_workbook(str(R_PATH / '参考面板3.5B.xlsx'), data_only=True)
    sheet = wb.active

    result = {}
    char_result = []
    char_temp = ''
    skill_temp = ''
    skill_count = -1
    title = 0
    for row in range(9, 400):
        temp = {}
        char_name = sheet.cell(row, 1).value
        skill = sheet.cell(row, 19).value
        if char_name and char_name != '角色' and isinstance(char_name, str):
            weapon = str(sheet.cell(row, 2).value)
            equip_set = str(sheet.cell(row, 3).value)
            if '4' in equip_set:
                equip_set = equip_set.replace('4', '')
                equip_set = await getEquipName(equip_set)
            elif '2' in equip_set:
                equip_set_list = equip_set[1:].split('2')
                equip_set = ''
                for i in equip_set_list:
                    equip_set += await getEquipName(i)
            else:
                print('error')

            equip_main = str(sheet.cell(row, 4).value)
            g_atk = sheet.cell(row, 8).value
            other = {}
            for i in [9, 10]:
                if sheet.cell(title, i).value is not None:
                    n = str(sheet.cell(title, i).value)
                    n = n.replace('力', '')
                    if '加成' in n:
                        continue
                    if sheet.cell(row, i).value is not None:
                        v = float(str(sheet.cell(row, i).value))
                        if v:
                            other[n] = v
            crit_rate = sheet.cell(row, 13).value
            crit_dmg = sheet.cell(row, 14).value

            weapon = weapon.replace('试做', '试作')
            char_name = (
                char_name.replace('空/荧', '旅行者')
                .replace('（', '(')
                .replace('）', ')')
            )
            if char_name != char_temp:
                skill_temp = ''
                skill_count = -1
            if char_name in SAMPLE:
                if isinstance(SAMPLE[char_name], str):
                    temp['skill'] = SAMPLE[char_name]
                else:
                    if skill != skill_temp:
                        skill_count += 1
                    skill_temp = skill
                    temp['skill'] = SAMPLE[char_name][skill_count]
                value = (
                    str(sheet.cell(row, 20).value)
                    if sheet.cell(row, 20).value
                    else ''
                )
                if '.' in value:
                    value = float(value)
                else:
                    value = 0
                temp['value'] = value
            else:
                temp['skill'] = ''
            temp['seq'] = '{}|{}|{}'.format(weapon, equip_set, equip_main)
            temp['critRate'] = crit_rate
            temp['critDmg'] = crit_dmg
            temp['atk'] = g_atk
            temp['other'] = other

            if char_temp:
                if char_name == char_temp:
                    pass
                else:
                    result[char_temp] = char_result
                    char_result = []
                    char_temp = char_name
            else:
                char_temp = char_name
            char_result.append(temp)
            if row == 354:
                print('ok!')
                result[char_temp] = char_result
        else:
            title = row
    with open(ETC_PATH / 'dmg_map.json', 'w', encoding='UTF-8') as file:
        json.dump(result, file, ensure_ascii=False)


async def main():
    await panle2Json()


asyncio.run(main())
