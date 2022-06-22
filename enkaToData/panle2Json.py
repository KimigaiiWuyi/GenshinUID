import asyncio
import json
import openpyxl
from pathlib import Path

import httpx

version = '2.7.0'
version_old = '2.6.0'

R_PATH = Path(__file__).parents[0]
DATA_PATH = R_PATH / 'data'
ETC_PATH = R_PATH / 'etc'

async def getEquipName(name: str) -> str:
    r = json.loads(httpx.get('https://info.minigg.cn/artifacts?query={}'.format(name)).text)
    re = r['name']
    return re

async def panle2Json() -> None:
    wb=openpyxl.load_workbook(DATA_PATH / '参考面板2.7（上）.xlsx',data_only = True)
    sheet=wb.active

    result = {}
    char_result = []
    char_temp = ''
    for row in range(9,300):
        temp = {}
        char_name = sheet.cell(row, 1).value
        if char_name and char_name != '角色':
            weapon = sheet.cell(row, 2).value
            equip_set = sheet.cell(row, 3).value
            if '4' in equip_set:
                equip_set = equip_set.replace('4','')
                equip_set = await getEquipName(equip_set)
            elif '2' in equip_set:
                equip_set_list = equip_set[1:].split('2')
                equip_set = ''
                for i in equip_set_list:
                    equip_set += await getEquipName(i)
            else:
                print('error')

            equip_main = sheet.cell(row, 4).value
            g_atk = sheet.cell(row, 8).value
            other = sheet.cell(row, 9).value
            other2 = sheet.cell(row, 10).value
            crit_rate = sheet.cell(row, 13).value

            if equip_main[1] in ['生']:
                key = '血量'
            elif equip_main[1] in ['精']:
                key = '元素精通'
            elif equip_main[1] in ['防']:
                key = '防御力'
            else:
                key = '攻击力'
            
            if char_name == '夜兰':
                key = '血量'
            elif char_name == '五郎':
                key = '防御力'
            

            dmgBouns = sheet.cell(row, 15).value
            defArea = sheet.cell(row, 16).value
            resArea = sheet.cell(row, 17).value
            power = sheet.cell(row, 18).value

            if char_name == '七七':
                power = '153+1174'
            elif power == '/' or power == 0:
                if char_name == '托马':
                    power = '14.40+4829'
                elif char_name == '班尼特':
                    power = '12.75+1587.82'
                elif char_name == '芭芭拉':
                    power = '35.2+4335'
                elif char_name == '早柚':
                    power = '159.74+1280'
                elif char_name == '琴':
                    power = '452.16+3389'
                elif char_name == '申鹤':
                    power = '攻击力'
                elif char_name == '五郎':
                    power = '防御力'
                elif char_name == '云堇':
                    power = '防御力'
                else:
                    power = 'any'

            action = sheet.cell(row, 19).value
            if sheet.cell(row, 20).value != 'any':
                val = float('{:.2f}'.format(float(sheet.cell(row, 20).value)))
            else:
                val = 'any'
            
            if char_name == '辛焱' and '盾' in action:
                power = '2.88 + 1773'

            #temp['name'] = char_name
            weapon = weapon.replace('试做','试作')
            temp['seq'] = '{}|{}|{}'.format(weapon, equip_set, equip_main)
            temp['action'] = action
            temp['crit_rate'] = crit_rate
            temp['atk'] = g_atk
            temp['dmgBouns'] = dmgBouns
            temp['defArea'] = defArea
            temp['resArea'] = resArea
            temp['other'] = other
            temp['other2'] = other2
            temp['key'] = key
            temp['power'] = power
            temp['val'] = val
            if char_temp:
                if char_name == char_temp:
                    pass
                else:
                    result[char_temp] = char_result
                    char_result = []
                    char_temp = char_name
            else:
                char_temp = char_name
            char_result.append(temp)
            if row == 263:
                print('ok!')
                result[char_temp] = char_result
    with open(DATA_PATH / 'dmgMap.json', 'w', encoding='UTF-8') as file:
        json.dump(result, file, ensure_ascii=False)

async def main():
    await panle2Json()

asyncio.run(main())