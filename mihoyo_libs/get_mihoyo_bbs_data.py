import math
import os
import sys
from base64 import b64encode
from io import BytesIO

from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from .get_data import *
from .get_image import draw_event_pic
import get_mihoyo_bbs_coin as coin

FILE_PATH = os.path.dirname(__file__)
FILE2_PATH = os.path.join(FILE_PATH, 'mihoyo_libs/mihoyo_bbs')
INDEX_PATH = os.path.join(FILE2_PATH, 'index')
Texture_PATH = os.path.join(FILE2_PATH, 'texture2d')

avatar_json = {
    "Albedo"   : "阿贝多",
    "Ambor"    : "安柏",
    "Barbara"  : "芭芭拉",
    "Beidou"   : "北斗",
    "Bennett"  : "班尼特",
    "Chongyun" : "重云",
    "Diluc"    : "迪卢克",
    "Diona"    : "迪奥娜",
    "Eula"     : "优菈",
    "Fischl"   : "菲谢尔",
    "Ganyu"    : "甘雨",
    "Hutao"    : "胡桃",
    "Jean"     : "琴",
    "Kazuha"   : "枫原万叶",
    "Kaeya"    : "凯亚",
    "Ayaka"    : "神里绫华",
    "Keqing"   : "刻晴",
    "Klee"     : "可莉",
    "Lisa"     : "丽莎",
    "Mona"     : "莫娜",
    "Ningguang": "凝光",
    "Noel"     : "诺艾尔",
    "Qiqi"     : "七七",
    "Razor"    : "雷泽",
    "Rosaria"  : "罗莎莉亚",
    "Sucrose"  : "砂糖",
    "Tartaglia": "达达利亚",
    "Venti"    : "温迪",
    "Xiangling": "香菱",
    "Xiao"     : "魈",
    "Xingqiu"  : "行秋",
    "Xinyan"   : "辛焱",
    "Yanfei"   : "烟绯",
    "Zhongli"  : "钟离",
    "Yoimiya"  : "宵宫",
    "Sayu"     : "早柚",
    "Shogun"   : "雷电将军",
    "Aloy"     : "埃洛伊",
    "Sara"     : "九条裟罗",
    "Kokomi"   : "珊瑚宫心海",
    "Shenhe"   : "申鹤"
}

daily_im = '''
*数据刷新可能存在一定延迟，请以当前游戏实际数据为准{}
==============
原粹树脂：{}/{}{}
每日委托：{}/{} 奖励{}领取
减半已用：{}/{}
洞天宝钱：{}
探索派遣：
总数/完成/上限：{}/{}/{}
{}'''

month_im = '''
==============
{}
UID：{}
==============
本日获取原石：{}
本日获取摩拉：{}
==============
昨日获取原石：{}
昨日获取摩拉：{}
==============
本月获取原石：{}
本月获取摩拉：{}
==============
上月获取原石：{}
上月获取摩拉：{}
==============
原石收入组成：
{}=============='''

weapon_im = '''【名称】：{}
【类型】：{}
【稀有度】：{}
【介绍】：{}
【攻击力】：{}{}{}'''

char_info_im = '''{}
【稀有度】：{}
【武器】：{}
【元素】：{}
【突破加成】：{}
【生日】：{}
【命之座】：{}
【cv】：{}
【介绍】：{}'''

artifacts_im = '''【{}】
【稀有度】：{}
【2件套】：{}
【4件套】：{}
【{}】：{}
【{}】：{}
【{}】：{}
【{}】：{}
【{}】：{}'''

food_im = '''【{}】
【稀有度】：{}
【食物类型】：{}
【食物类别】：{}
【效果】：{}
【介绍】：{}
【材料】：
{}'''

audio_json = '''{
    "357":["357_01","357_02","357_03"],
    "1000000":["1000000_01","1000000_02","1000000_03","1000000_04","1000000_05","1000000_06","1000000_07"],
    "1000001":["1000001_01","1000001_02","1000001_03"],
    "1000002":["1000002_01","1000002_02","1000002_03"],
    "1000100":["1000100_01","1000100_02","1000100_03","1000100_04","1000100_05"],
    "1000101":["1000101_01","1000101_02","1000101_03","1000101_04","1000101_05","1000101_06"],
    "1000200":["1000200_01","1000200_02","1000200_03"],
    "1010201":["1010201_01"],
    "1000300":["1000300_01","1000300_02"],
    "1000400":["1000400_01","1000400_02","1000400_03"],
    "1000500":["1000500_01","1000500_02","1000500_03"],
    "1010000":["1010000_01","1010000_02","1010000_03","1010000_04","1010000_05"],
    "1010001":["1010001_01","1010001_02"],
    "1010100":["1010100_01","1010100_02","1010100_03","1010100_04","1010100_05"],
    "1010200":["1010200_01","1010200_02","1010200_03","1010200_04","1010200_05"],
    "1010300":["1010300_01","1010300_02","1010300_03","1010300_04","1010300_05"],
    "1010301":["1010301_01","1010301_02","1010301_03","1010301_04","1010301_05"],
    "1010400":["1010400_01","1010400_02","1010400_03"],
    "1020000":["1020000_01"]
}'''

char_adv_im = '''【{}】
【五星武器】：{}
【四星武器】：{}
【三星武器】：{}
【圣遗物】：
{}'''


async def weapon_adv(name):
    char_adv_path = os.path.join(FILE_PATH, "mihoyo_libs/Genshin All Char.xlsx")
    wb = load_workbook(char_adv_path)
    ws = wb.active

    weapon_name = ""
    char_list = []
    for c in range(2, 5):
        for r in range(2, 300):
            if ws.cell(r, c).value:
                # if all(i in ws.cell(r,c).value for i in name):
                if name in ws.cell(r, c).value:
                    weapon_name = ws.cell(r, c).value
                    char_list.append(ws.cell(2 + ((r - 2) // 5) * 5, 1).value)

    if char_list:
        im = ','.join(char_list)
        im = im + "可能会用到【{}】".format(weapon_name)
    else:
        im = "没有角色能使用【{}】".format(weapon_name)
    return im


async def char_adv(name):
    char_name = None
    char_adv_path = os.path.join(FILE_PATH, "mihoyo_libs/Genshin All Char.xlsx")
    wb = load_workbook(char_adv_path)
    ws = wb.active
    char_list = ws["A"]
    index = None
    for i in char_list:
        if i.value:
            if all(g in i.value for g in name):
                # if name in i.value:
                index = i.row
                char_name = i.value
    if index:
        weapon_5star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 2).value:
                weapon_5star += ws.cell(i, 2).value + ">"
        if weapon_5star != "":
            weapon_5star = weapon_5star[:-1]
        else:
            weapon_5star = "无推荐"

        weapon_4star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 3).value:
                weapon_4star += ws.cell(i, 3).value + ">"
        if weapon_4star != "":
            weapon_4star = weapon_4star[:-1]
        else:
            weapon_4star = "无推荐"

        weapon_3star = ""
        for i in range(index, index + 5):
            if ws.cell(i, 4).value:
                weapon_3star += ws.cell(i, 4).value + ">"
        if weapon_3star != "":
            weapon_3star = weapon_3star[:-1]
        else:
            weapon_3star = "无推荐"

        artifacts = ""
        for i in range(index, index + 5):
            if ws.cell(i, 5).value:
                if ws.cell(i, 6).value:
                    artifacts += ws.cell(i, 5).value + "*2" + ws.cell(i, 6).value + "*2" + "\n"
                else:
                    artifacts += ws.cell(i, 5).value + "*4" + "\n"

        if artifacts != "":
            artifacts = artifacts[:-1]
        else:
            artifacts = "无推荐"

        im = char_adv_im.format(char_name, weapon_5star, weapon_4star, weapon_3star, artifacts)
        return im


async def deal_ck(mes, qid):
    if "stoken" in mes:
        login_ticket = re.search(r"login_ticket=([0-9a-zA-Z]+)", mes).group(0).split('=')[1]
        uid = await select_db(qid, "uid")
        # mys_id = re.search(r"login_uid=([0-9]+)", mes).group(0).split('=')[1]
        ck = await owner_cookies(uid[0])
        mys_id = re.search(r"account_id=(\d*)", ck).group(0).split('=')[1]
        raw_data = await get_stoken_by_login_ticket(login_ticket, mys_id)
        stoken = raw_data["data"]["list"][0]["token"]
        s_cookies = "stuid={};stoken={}".format(mys_id, stoken)
        await stoken_db(s_cookies, uid[0])
        return "添加Stoken成功！"
    else:
        aid = re.search(r"account_id=(\d*)", mes)
        mysid_data = aid.group(0).split('=')
        mysid = mysid_data[1]
        cookie = ';'.join(filter(lambda x: x.split('=')[0] in [
            "cookie_token", "account_id"], [i.strip() for i in mes.split(';')]))
        mys_data = await get_mihoyo_bbs_info(mysid, cookie)
        for i in mys_data['data']['list']:
            if i['game_id'] != 2:
                mys_data['data']['list'].remove(i)
        uid = mys_data['data']['list'][0]['game_role_id']

        conn = sqlite3.connect('ID_DATA.db')
        c = conn.cursor()

        try:
            c.execute("DELETE from CookiesCache where uid=? or mysid = ?", (uid, mysid))
        except:
            pass

        conn.commit()
        conn.close()

        await cookies_db(uid, cookie, qid)
        return f'添加Cookies成功！\nCookies属于个人重要信息，如果你是在不知情的情况下添加，请马上修改米游社账户密码，保护个人隐私！\n————\n' \
               f'如果需要【开启自动签到】和【开启推送】还需要在【群聊中】使用命令“绑定uid”绑定你的uid。\n例如：绑定uid123456789。'


async def award(uid):
    data = await get_award(uid)
    nickname = data['data']['nickname']
    day_stone = data['data']['day_data']['current_primogems']
    day_mora = data['data']['day_data']['current_mora']
    lastday_stone = data['data']['day_data']['last_primogems']
    lastday_mora = data['data']['day_data']['last_mora']
    month_stone = data['data']['month_data']['current_primogems']
    month_mora = data['data']['month_data']['current_mora']
    lastmonth_stone = data['data']['month_data']['last_primogems']
    lastmonth_mora = data['data']['month_data']['last_mora']
    group_str = ''
    for i in data['data']['month_data']['group_by']:
        group_str = group_str + \
                    i['action'] + "：" + str(i['num']) + \
                    "（" + str(i['percent']) + "%）" + '\n'

    im = month_im.format(nickname, uid, day_stone, day_mora, lastday_stone, lastday_mora,
                         month_stone, month_mora, lastmonth_stone, lastmonth_mora, group_str)
    return im


async def audio_wiki(name, message):
    async def get(_audioid):
        tmp_json = json.loads(audio_json)
        for _ in range(3):  # 重试3次
            if _audioid in tmp_json:
                if not tmp_json[_audioid]:
                    return
                audioid1 = random.choice(tmp_json[_audioid])
            else:
                audioid1 = _audioid
            url = await get_audio_info(name, audioid1)
            req = requests.get(url)
            if req.status_code == 200:
                return BytesIO(req.content)
            else:
                if _audioid in tmp_json:
                    tmp_json[_audioid].remove(audioid1)

    if name == "列表":
        imgmes = 'base64://' + b64encode(open(os.path.join(INDEX_PATH, "语音.png"), "rb").read()).decode()
        return imgmes
    elif name == "":
        return "请输入角色名。"
    else:
        audioid = re.findall(r"[0-9]+", message)[0]
        try:
            audio = await get(audioid)
        except:
            return "语音获取失败"
        if audio:
            audios = 'base64://' + b64encode(audio.getvalue()).decode()
            return audios
        else:
            return "没有找到语音，请检查语音ID与角色名是否正确，如无误则可能未收录该语音"


async def artifacts_wiki(name):
    data = await get_misc_info("artifacts", name)
    if "errcode" in data:
        im = "该圣遗物不存在。"
    else:
        star = ""
        for i in data["rarity"]:
            star = star + i + "星、"
        star = star[:-1]
        im = artifacts_im.format(data["name"], star, data["2pc"], data["4pc"], data["flower"]["name"],
                                 data["flower"]["description"],
                                 data["plume"]["name"], data["plume"]["description"], data["sands"]["name"],
                                 data["sands"]["description"],
                                 data["goblet"]["name"], data["goblet"]["description"], data["circlet"]["name"],
                                 data["circlet"]["description"])
    return im


async def foods_wiki(name):
    data = await get_misc_info("foods", name)
    if "errcode" in data:
        im = "该食物不存在。"
    else:
        ingredients = ""
        food_temp = {}
        for i in data["ingredients"]:
            if i["name"] not in food_temp:
                food_temp[i["name"]] = i["count"]
            else:
                food_temp[i["name"]] = food_temp[i["name"]] + i["count"]
        for i in food_temp:
            ingredients += i + ":" + str(food_temp[i]) + "\n"
        ingredients = ingredients[:-1]
        im = food_im.format(data["name"], data["rarity"], data["foodtype"], data["foodfilter"], data["effect"],
                            data["description"], ingredients)
    return im


async def enemies_wiki(name):
    raw_data = await get_misc_info("enemies", name)
    if "errcode" in raw_data:
        im = "该原魔不存在。"
    else:
        reward = ""
        for i in raw_data["rewardpreview"]:
            reward += i["name"] + "：" + str(i["count"]) if "count" in i.keys() else i["name"] + "：" + "可能"
            reward += "\n"
        im = "【{}】\n——{}——\n类型：{}\n信息：{}\n掉落物：\n{}".format(raw_data["name"], raw_data["specialname"],
                                                           raw_data["category"], raw_data["description"], reward)
    return im


# 签到函数
async def sign(uid):
    try:
        sign_data = await mihoyo_bbs_sign(uid)
        sign_info = await get_sign_info(uid)
        sign_info = sign_info['data']
        sign_list = await get_sign_list()
        status = sign_data['message']
        getitem = sign_list['data']['awards'][int(
            sign_info['total_sign_day']) - 1]['name']
        getnum = sign_list['data']['awards'][int(
            sign_info['total_sign_day']) - 1]['cnt']
        get_im = f"本次签到获得{getitem}x{getnum}"
        if status == "OK" and sign_info['is_sign'] == True:
            mes_im = "签到成功"
        else:
            mes_im = status
        sign_missed = sign_info['sign_cnt_missed']
        im = mes_im + "!" + "\n" + get_im + "\n" + f"本月漏签次数：{sign_missed}"
    except:
        im = "签到失败，请检查Cookies是否失效。"
    return im


# 统计状态函数
async def daily(mode="push", uid=None):
    def seconds2hours(seconds: int) -> str:
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        return "%02d:%02d:%02d" % (h, m, s)

    temp_list = []
    conn = sqlite3.connect('ID_DATA.db')
    c = conn.cursor()
    c_data = None
    if mode == "push":
        cursor = c.execute(
            "SELECT *  FROM NewCookiesTable WHERE StatusA != ?", ("off",))
        c_data = cursor.fetchall()
    elif mode == "ask":
        c_data = ([uid, 0, 0, 0, 0, 0, 0],)

    for row in c_data:
        raw_data = await get_daily_data(str(row[0]))
        if raw_data["retcode"] != 0:
            temp_list.append(
                {"qid": row[2], "gid": row[3], "message": "你的推送状态有误；可能是uid绑定错误或没有在米游社打开“实时便筏”功能。"})
        else:
            dailydata = raw_data["data"]
            current_resin = dailydata['current_resin']

            current_expedition_num = dailydata['current_expedition_num']
            max_expedition_num = dailydata['max_expedition_num']
            finished_expedition_num = 0
            expedition_info: list[str] = []
            for expedition in dailydata['expeditions']:
                avatar: str = expedition['avatar_side_icon'][89:-4]
                try:
                    avatar_name: str = avatar_json[avatar]
                except KeyError:
                    avatar_name: str = avatar

                if expedition['status'] == 'Finished':
                    expedition_info.append(f"{avatar_name} 探索完成")
                    finished_expedition_num += 1
                else:
                    remained_timed: str = seconds2hours(
                        expedition['remained_time'])
                    expedition_info.append(
                        f"{avatar_name} 剩余时间{remained_timed}")

            # 推送条件检查，在指令查询时 row[6] 为 0 ，而自动推送时 row[6] 为 140，这样保证用指令查询时必回复
            # 说实话我仔细看了一会才理解…
            if current_resin >= row[6] or dailydata["max_home_coin"] - dailydata["current_home_coin"] <= 100:
                tip = ''
                tips = []
                if current_resin >= row[6] != 0:
                    tips.append("你的树脂快满了！")
                if dailydata["max_home_coin"] - dailydata["current_home_coin"] <= 100:
                    tips.append("你的洞天宝钱快满了！")
                if finished_expedition_num == current_expedition_num:
                    tips.append("你的所有探索派遣完成了！")  # emmmm
                if tips:
                    tips.insert(0, '\n==============')
                    tip = '\n'.join(tips)

                max_resin = dailydata['max_resin']
                rec_time = ''
                # logger.info(dailydata)
                if current_resin < 160:
                    resin_recovery_time = seconds2hours(
                        dailydata['resin_recovery_time'])
                    next_resin_rec_time = seconds2hours(
                        8 * 60 - ((dailydata['max_resin'] - dailydata['current_resin']) * 8 * 60 - int(
                            dailydata['resin_recovery_time'])))
                    rec_time = f' ({next_resin_rec_time}/{resin_recovery_time})'

                finished_task_num = dailydata['finished_task_num']
                total_task_num = dailydata['total_task_num']
                is_extra_got = '已' if dailydata['is_extra_task_reward_received'] else '未'

                resin_discount_num_limit = dailydata['resin_discount_num_limit']
                used_resin_discount_num = resin_discount_num_limit - \
                                          dailydata['remain_resin_discount_num']

                home_coin = f'{dailydata["current_home_coin"]}/{dailydata["max_home_coin"]}'
                if dailydata["current_home_coin"] < dailydata["max_home_coin"]:
                    coin_rec_time = seconds2hours(int(dailydata["home_coin_recovery_time"]))
                    coin_add_speed = math.ceil((dailydata["max_home_coin"] - dailydata["current_home_coin"]) / (
                            int(dailydata["home_coin_recovery_time"]) / 60 / 60))
                    home_coin += f'（{coin_rec_time} 约{coin_add_speed}/h）'

                expedition_data = "\n".join(expedition_info)
                send_mes = daily_im.format(tip, current_resin, max_resin, rec_time, finished_task_num, total_task_num,
                                           is_extra_got, used_resin_discount_num,
                                           resin_discount_num_limit, home_coin, current_expedition_num,
                                           finished_expedition_num, max_expedition_num, expedition_data)

                temp_list.append(
                    {"qid": row[2], "gid": row[3], "message": send_mes})
    return temp_list


async def mihoyo_coin(qid, s_cookies=None):
    uid = await select_db(qid, mode="uid")
    uid = uid[0]
    if s_cookies is None:
        s_cookies = await get_stoken(uid)

    if s_cookies:
        get_coin = coin.mihoyobbs_coin(s_cookies)
        im = await get_coin.task_run()
    else:
        im = "你还没有绑定Stoken~"
    return im


async def get_event_pic():
    img_path = os.path.join(FILE2_PATH, "event.jpg")
    while True:
        if os.path.exists(img_path):
            f = open(img_path, 'rb')
            ls_f = b64encode(f.read()).decode()
            img_mes = 'base64://' + ls_f
            f.close()
            break
        else:
            await draw_event_pic()
    return img_mes


async def weapon_wiki(name, level=None):
    data = await get_weapon_info(name)
    if "errcode" in data:
        im = "武器不存在。"
    elif level:
        data2 = await get_weapon_info(name, level)
        if data["substat"] != "":
            sp = data["substat"] + "：" + '%.1f%%' % (data2["specialized"] * 100) \
                if data["substat"] != "元素精通" else data["substat"] + "：" + str(math.floor(data2["specialized"]))
        else:
            sp = ""
        im = (data["name"] + "\n等级：" + str(data2["level"]) + "（突破" + str(data2["ascension"]) + "）" +
              "\n攻击力：" + str(round(data2["attack"])) + "\n" + sp)
    else:
        name = data['name']
        _type = data['weapontype']
        star = data['rarity'] + "星"
        info = data['description']
        atk = str(data['baseatk'])
        sub_name = data['substat']
        if data['subvalue'] != "":
            sub_val = (data['subvalue'] +
                       '%') if sub_name != '元素精通' else data['subvalue']
            sub = "\n" + "【" + sub_name + "】" + sub_val
        else:
            sub = ""

        if data['effectname'] != "":
            raw_effect = data['effect']
            rw_ef = []
            for i in range(len(data['r1'])):
                now = ''
                for j in range(1, 6):
                    now = now + data['r{}'.format(j)][i] + "/"
                now = now[:-1]
                rw_ef.append(now)
            raw_effect = raw_effect.format(*rw_ef)
            effect = "\n" + "【" + data['effectname'] + "】" + "：" + raw_effect
        else:
            effect = ""
        im = weapon_im.format(name, _type, star, info, atk,
                              sub, effect)
    return im


async def char_wiki(name, mode="char", level=None):
    im = ''
    data = await get_char_info(name, mode, level if mode == "char" else None)
    if mode == "char":
        if isinstance(data, list):
            im = ','.join(data)
        elif "errcode" in data:
            im = "不存在该角色或类型。"
        elif level:
            data2 = await get_char_info(name, mode)
            sp = data2["substat"] + "：" + '%.1f%%' % (data["specialized"] * 100) if data2["substat"] != "元素精通" else \
                data2["substat"] + "：" + str(math.floor(data["specialized"]))
            im = (data2["name"] + "\n等级：" + str(data["level"]) + "\n血量：" + str(math.floor(data["hp"])) +
                  "\n攻击力：" + str(math.floor(data["attack"])) + "\n防御力：" + str(math.floor(data["defense"])) +
                  "\n" + sp)
        else:
            name = data['title'] + ' — ' + data['name']
            star = data['rarity']
            _type = data["weapontype"]
            element = data['element']
            up_val = data['substat']
            bdday = data['birthday']
            polar = data['constellation']
            cv = data['cv']['chinese']
            info = data['description']
            im = char_info_im.format(
                name, star, _type, element, up_val, bdday, polar, cv, info)
    elif mode == "costs":
        if isinstance(data[1], list):
            im = ','.join(data[1])
        elif "errcode" in data[1]:
            im = "不存在该角色或类型。"
        else:
            im = "【天赋材料(一份)】\n{}\n【突破材料】\n{}"
            im1 = ""
            im2 = ""

            talent_temp = {}
            talent_cost = data[1]
            for i in talent_cost.values():
                for j in i:
                    if j["name"] not in talent_temp:
                        talent_temp[j["name"]] = j["count"]
                    else:
                        talent_temp[j["name"]] = talent_temp[j["name"]] + j["count"]
            for k in talent_temp:
                im1 = im1 + k + ":" + str(talent_temp[k]) + "\n"

            temp = {}
            cost = data[0]
            for i in range(1, 7):
                for j in cost["ascend{}".format(i)]:
                    if j["name"] not in temp:
                        temp[j["name"]] = j["count"]
                    else:
                        temp[j["name"]] = temp[j["name"]] + j["count"]

            for k in temp:
                im2 = im2 + k + ":" + str(temp[k]) + "\n"

            im = im.format(im1, im2)
    elif mode == "constellations":
        if "errcode" in data:
            im = "不存在该角色或命座数量。"
        else:
            im = "【" + data["c{}".format(level)]['name'] + "】" + "：" + \
                 "\n" + data["c{}".format(level)]['effect'].replace("*", "")
    elif mode == "talents":
        if "errcode" in data:
            im = "不存在该角色。"
        else:
            if 7 >= int(level) > 0:
                if int(level) <= 3:
                    if level == "1":
                        data = data["combat1"]
                    elif level == "2":
                        data = data["combat2"]
                    elif level == "3":
                        data = data["combat3"]
                    skill_name = data["name"]
                    skill_info = data["info"]
                    skill_detail = ""

                    """
                    for i in data["attributes"]["parameters"]:
                        temp = ""
                        for k in data["attributes"]["parameters"][i]:
                            if str(k).count('.') == 1:
                                temp += "%.2f%%" % (k * 100) + "/"
                            else:
                                temp = k
                                break
                        data["attributes"]["parameters"][i] = temp[:-1]

                    for i in data["attributes"]["labels"]:
                        #i = i.replace("{","{{")
                        i = re.sub(r':[a-zA-Z0-9]+}', "}", i)
                        #i.replace(r':[a-zA-Z0-9]+}','}')
                        skill_detail += i + "\n"

                    skill_detail = skill_detail.format(**data["attributes"]["parameters"])
                    """
                    mes_list = []
                    parameters = []
                    add_switch = True
                    for i in data["attributes"]["parameters"]:
                        for index, j in enumerate(data["attributes"]["parameters"][i]):
                            if add_switch:
                                parameters.append({})
                            if str(j).count('.') == 1 and j <= 20:
                                parameters[index].update({i: "%.2f%%" % (j * 100)})
                            elif str(j).count('.') == 1:
                                parameters[index].update({i: "%.2f" % (j * 100)})
                            else:
                                parameters[index].update({i: j})
                        add_switch = False

                    for k in data["attributes"]["labels"]:
                        k = re.sub(r':[a-zA-Z0-9]+}', "}", k)
                        skill_detail += k + "\n"

                    skill_detail = skill_detail[:-1]

                    for i in range(1, 10):
                        if i % 2 != 0:
                            skill_info = skill_info.replace("**", "「", 1)
                        else:
                            skill_info = skill_info.replace("**", "」", 1)

                    mes_list.append({
                        "type": "node",
                        "data": {
                            "name"   : "小仙",
                            "uin"    : "3399214199",
                            "content": "【" + skill_name + "】" + "\n" + skill_info
                        }
                    })

                    for index, i in enumerate(parameters):
                        mes = skill_detail.format(**i)
                        node_data = {
                            "type": "node",
                            "data": {
                                "name"   : "小仙",
                                "uin"    : "3399214199",
                                "content": "lv." + str(index + 1) + "\n" + mes
                            }
                        }
                        mes_list.append(node_data)
                    im = mes_list

                else:
                    if level == "4":
                        data = data["passive1"]
                    elif level == "5":
                        data = data["passive2"]
                    elif level == "6":
                        data = data["passive3"]
                    elif level == "7":
                        data = data["passive4"]
                    skill_name = data["name"]
                    skill_info = data["info"]
                    im = "【{}】\n{}".format(skill_name, skill_info)
            else:
                im = "不存在该天赋。"
    return im