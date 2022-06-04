import copy
import json
from typing import Optional

from openpyxl import load_workbook

sample = {
    'weapon': {
        '5': [],
        '4': [],
        '3': []
    },
    'artifact': [],  # type: list[list[str,Optional[str]]]  # 四件套 / 2+2
    'remark': []
}

char_json = {}

wb = load_workbook('../genshin_uid/mihoyo_libs/Genshin All Char.xlsx')
ws = wb.active
for char_i in range(2, 336, 5):  # 角色行
    char_name = ws.cell(char_i, 1).value.replace('\n', '')
    char_sample = copy.deepcopy(sample)
    for i in range(5):
        row = i + char_i

        if star_5 := ws.cell(row, 2).value:
            char_sample['weapon']['5'].append(star_5)
        if star_4 := ws.cell(row, 3).value:
            char_sample['weapon']['4'].append(star_4)

        artifact = []
        if arti_1 := ws.cell(row, 5).value:
            artifact.append(arti_1)
        if arti_2 := ws.cell(row, 6).value:
            artifact.append(arti_2)
        if artifact:
            char_sample['artifact'].append(artifact)

        if remark := ws.cell(row, 7).value:
            if row > 7:
                char_sample['remark'].append(remark)

    char_json[char_name] = char_sample

with open('../genshin_uid/mihoyo_libs/char_adv_list.json', 'w', encoding='utf-8') as f:
    json.dump(char_json, f, indent=2, ensure_ascii=False)