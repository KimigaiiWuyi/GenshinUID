import json
import asyncio
from pathlib import Path

import httpx
import openpyxl

R_PATH = Path(__file__).parent
DATA_PATH = R_PATH / 'blue_data'
ETC_PATH = Path(__file__).parents[1] / 'genshinuid_enka' / 'dmgCalc'


async def get_misc_info(mode: str, name: str):
    """
    :说明:
      一些杂项信息。
    :参数:
      * name (str): 'enemies', 'foods', 'artifacts'。
      * name (str): 参数。
    :返回:
      * data (str): 获取数据信息。
    """
    url = 'https://info.minigg.cn/{}'.format(mode)
    req = httpx.get(url=url, params={'query': name})
    return req.json()


async def getEquipName(name: str) -> str:
    r = await get_misc_info('artifacts', name)
    re = r['name']
    return re


async def panle2Json() -> None:
    """
    :说明:
      访问DATA_PATH并转换数据为dmgMap.json。
    """
    wb = openpyxl.load_workbook(
        str(DATA_PATH / '参考面板3.0.xlsx'), data_only=True
    )
    sheet = wb.active

    result = {}
    char_result = []
    char_temp = ''
    title = 0
    for row in range(9, 300):
        temp = {}
        char_name = sheet.cell(row, 1).value
        if char_name and char_name != '角色':
            weapon = str(sheet.cell(row, 2).value)
            equip_set = str(sheet.cell(row, 3).value)
            if '4' in equip_set:
                equip_set = equip_set.replace('4', '')
                equip_set = await getEquipName(equip_set)
            elif '2' in equip_set:
                equip_set_list = equip_set[1:].split('2')
                equip_set = ''
                for i in equip_set_list:
                    equip_set += await getEquipName(i)
            else:
                print('error')

            equip_main = str(sheet.cell(row, 4).value)
            g_atk = sheet.cell(row, 8).value
            other = {}
            for i in [9, 10]:
                if sheet.cell(title, i).value is not None:
                    n = str(sheet.cell(title, i).value)
                    if '加成' in n:
                        continue
                    if sheet.cell(row, i).value is not None:
                        v = float(str(sheet.cell(row, i).value))
                        if v:
                            other[sheet.cell(title, i).value] = v
            crit_rate = sheet.cell(row, 13).value
            crit_dmg = sheet.cell(row, 14).value

            weapon = weapon.replace('试做', '试作')
            temp['seq'] = '{}|{}|{}'.format(weapon, equip_set, equip_main)
            temp['critRate'] = crit_rate
            temp['critDmg'] = crit_dmg
            temp['atk'] = g_atk
            temp['other'] = other

            if char_temp:
                if char_name == char_temp:
                    pass
                else:
                    result[char_temp] = char_result
                    char_result = []
                    char_temp = char_name
            else:
                char_temp = char_name
            char_result.append(temp)
            if row == 296:
                print('ok!')
                result[char_temp] = char_result
        else:
            title = row
    result['旅行者'] = [
        {
            "seq": "护摩之杖|无锋剑|生火暴",
            "critRate": 0.65,
            "critDmg": 1.55,
            "atk": 2300,
            "other": {"生命": 16000, "元素精通": 45},
        }
    ]
    with open(ETC_PATH / 'dmgMap.json', 'w', encoding='UTF-8') as file:
        json.dump(result, file, ensure_ascii=False)


async def main():
    await panle2Json()


asyncio.run(main())
